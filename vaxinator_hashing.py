from tkinter import *
from tkinter import messagebox
from tkinter import ttk  # themed widgets
from PIL import Image, ImageTk  # loading images
from openpyxl import load_workbook
from datetime import date, timedelta
import datetime



class Hashing:
    # to find key
    def Hash_Value(self, key):
        # print(key,type(key))
        self.hash_value = key % 43
        if self.hash_value == 0:
            self.hash_value = 1
        return self.hash_value

    # to avoid collission
    def Linear_Probing(self, curr_idx, records):
        new_index = None
        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active

        for forward_idx in range(curr_idx, 43):

            cell_chk2 = sheet.cell(row=forward_idx, column=2).value
            if cell_chk2 is None:
                for col in range(1, 9):
                    cellref = sheet.cell(row=forward_idx, column=col)
                    cellref.value = records[col - 1]

                new_index = forward_idx
                break

            elif forward_idx == 42:
                for backward_idx in range(1, curr_idx):
                    cell_chk2 = sheet.cell(row=backward_idx, column=2).value

                    if cell_chk2 is None:
                        for col in range(1, 9):
                            cellref = sheet.cell(row=backward_idx, column=col)
                            cellref.value = records[col - 1]

                        new_index = backward_idx
                        break

        if new_index is None:
            messagebox.showerror('ERROR!',
                                 'File is Full. Cannot Insert any Record\nPlease Delete Unwanted Records to make space')
            screen1.destroy()

        else:
            workbook.save(filename="record_info.xlsx")

            messagebox.showinfo('SUCCESS!', 'Record Successfully Inserted at position {}'.format(new_index))
            screen1.destroy()

    # insert records
    def Insert(self, key, values):
        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active
        result = False

        for c in sheet['B']:
            if c.value == key:
                result = True
                break
        if result:
            messagebox.showerror('DUPLICATE!', 'Aadhaar Number Already Exists!')
            adhaar.delete(0, END)  # delete only aadhaar number entered by user
            screen1.state('zoomed')

        else:

            cell_chk = sheet.cell(row=self.hash_value, column=2).value
            if cell_chk is not None:
                messagebox.showwarning('COLLISSION!',
                                       'Collission has occurred at position {} \nIt will be resolved by Linear Probing'.format(
                                           self.hash_value))
                screen1.destroy()
                self.Linear_Probing(self.hash_value, values)

            else:

                for col in range(1, 9):
                    cellref = sheet.cell(row=self.hash_value, column=col)
                    cellref.value = values[col - 1]

                workbook.save(filename="record_info.xlsx")
                messagebox.showinfo('SUCCESS!', 'Record Successfully Inserted at position {}'.format(self.hash_value))
                screen1.destroy()

    # search records
    def Search(self, sch_key):

        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active
        now = date.today()

        fnd = self.locate(self.hash_value, sch_key)

        if fnd:
            screen3.title("VAXINATOR - search [Record found at position {} ]".format(fnd))
            screen2.destroy()
            sheet = workbook.active
            d1 = sheet["F" + str(fnd)].value
            year1, month1, day1 = map(int, d1.split('-'))
            date1 = datetime.date(year1, month1, day1)
            date2 = sheet["G" + str(fnd)].value
            vac = sheet["E" + str(fnd)].value
            Label(screen3, font=('tahoma', 10), text="DETAILS:\n").pack()
            head = 0

            for data in sheet[str(fnd)]:
                ele = header[head]
                Label(screen3, font=('tahoma', 10), text=ele + str(data.value).upper()).pack()
                head += 1

            Label(screen3, text="\n").pack()

            if date2 == "na":
                diff = (now - date1).days
                Label(screen3, font=('tahoma', 10), text="Number of Days Completed Since 1st Dose:").pack()
                Label(screen3, fg='blue', font=('tahoma', 12), text=diff).pack()

                if vac == 'covishield':
                    max_days = 84
                    Label(screen3, font=('tahoma', 10), text="\nMinimum Interval for COVISHIELD:").pack()
                    Label(screen3, fg='green', font=('tahoma', 12), text='84 Days').pack()

                else:
                    max_days = 28
                    Label(screen3, font=('tahoma', 10), text="\nMinimum Interval for COVAXIN:").pack()
                    Label(screen3, fg='blue', font=('tahoma', 12), text='28 Days').pack()

                if diff >= max_days:
                    Label(screen3, fg='green', font=('tahoma', 12),
                          text="\nThe Person is ELIGIBLE for Second Dose").pack()

                else:
                    Label(screen3, fg='red', font=('tahoma', 12),
                          text="\nThe Person is NOT ELIGIBLE for Second Dose Now").pack()
                    visit = date1 + timedelta(days=max_days)
                    Label(screen3, font=('tahoma', 10), text="\nEligible for second dose from:").pack()
                    Label(screen3, fg='blue', font=('tahoma', 12), text=str(visit)).pack()

            else:
                Label(screen3, fg='green', font=('tahoma', 12), text="The Person is Completely Vaccinated!!").pack()


        else:
            messagebox.showerror('Unavailable', 'Record Not Found!')
            screen3.destroy()
            screen2.destroy()
        workbook.save(filename="record_info.xlsx")

    # display records
    def Display(self):

        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active

        # scroll bar
        main_frame = Frame(screen4)
        main_frame.pack(fill=BOTH, expand=1)
        sec = Frame(main_frame)
        sec.pack(fill=X, side=BOTTOM)
        my_canvas = Canvas(main_frame)
        my_canvas.pack(side=LEFT, fill=BOTH, expand=1)
        y_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
        y_scrollbar.pack(side=RIGHT, fill=Y)
        my_canvas.configure(yscrollcommand=y_scrollbar.set)
        my_canvas.bind("<Configure>", lambda e: my_canvas.config(scrollregion=my_canvas.bbox(ALL)))
        second_frame = Frame(my_canvas)
        my_canvas.create_window((0, 0), window=second_frame, anchor="n")
        hd = 0

        for row_idx, tup in enumerate(
                sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column, values_only=True)):

            if tup[0] is not None:
                row_num = row_idx + 1
                Label(second_frame, bg='white', fg='blue', font=('tahoma', 10),
                      text="POSITION OF RECORD: {}".format(row_num)).pack()
                for data in tup:
                    hvalue = header[hd]
                    Label(second_frame, bg='white', font=('tahoma', 10), text=hvalue + str(data).upper()).pack()
                    hd += 1

                hd = 0
                Label(second_frame, text='\n').pack()
        workbook.save(filename="record_info.xlsx")

    # update records
    def Update(self, idx_val):

        dose2_update = dose2upd.get()
        status_update = statusupd.get()
        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active
        sheet["G" + str(idx_val)].value = dose2_update
        sheet["H" + str(idx_val)].value = status_update
        workbook.save(filename="record_info.xlsx")
        messagebox.showinfo('SUCCESSS!', 'Record Updated Successfully!')
        screen6.destroy()

    # delete records
    def Remove(self, index):

        res = messagebox.askquestion('DELETE', 'Are you sure to Delete the Record?')
        if res == 'yes':
            workbook = load_workbook(filename="record_info.xlsx")
            sheet = workbook.active
            # sheet.delete_rows(idx=index)
            for col_num in range(1, 9):
                cellref = sheet.cell(row=index, column=col_num)
                cellref.value = None

            workbook.save(filename="record_info.xlsx")

            messagebox.showinfo('SUCCESS!', 'Record Deleted')
            screen8.destroy()
            screen7.destroy()

        else:
            screen8.destroy()

    # choice 1: Insert
    def getvalues():

        global screen1
        screen1 = Toplevel(screen)  # A Toplevel widget creates a screen on top of all other screens.
        screen1.title("VAXINATOR - insert")
        screen1.geometry("1080x600")

        global uname
        global adhaar
        global date_val
        global phone
        global vaccine
        global doseone
        global dosetwo
        global stat

        h = Hashing()

        uname = StringVar()
        adhaar = int
        date_val = StringVar()
        phone = int
        vaccine = StringVar()
        doseone = StringVar()
        dosetwo = StringVar()
        stat = StringVar()

        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Full Name').pack()
        uname = Entry(screen1, text=uname, relief='solid')
        uname.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Aadhar Number').pack()
        adhaar = Entry(screen1, text=adhaar, relief='solid')
        adhaar.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Date of Birth (yyyy-mm-dd)').pack()
        date_val = Entry(screen1, text=date_val, relief='solid')
        date_val.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Mobile Number').pack()
        phone = Entry(screen1, text=phone, relief='solid')
        phone.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Vaccine Name').pack()
        vaccine = Entry(screen1, text=vaccine, relief='solid')
        vaccine.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Dose 1 date (yyyy-mm-dd)').pack()
        doseone = Entry(screen1, text=doseone, relief='solid')
        doseone.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Dose 2 date (yyyy-mm-dd)').pack()
        dosetwo = Entry(screen1, text=dosetwo, relief='solid')
        dosetwo.pack()
        Label(screen1, height=2, width=50, font=('tahoma', 10), text='Status (partial/full)').pack()
        stat = Entry(screen1, text=stat, relief='solid')
        stat.pack()
        Label(text="").pack()
        Button(screen1, bg='#ebf6f9', fg='darkgreen', activebackground='green', activeforeground='white', text="SUBMIT",
               width=10, height=1, command=h.insert_values).pack()

    def locate(self, hv, key_adhr):

        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active
        fnd = None
        seq_srh = None
        for fwd_idx in range(hv, 43):
            seq_srh = sheet.cell(row=fwd_idx, column=2).value
            if seq_srh == key_adhr:
                fnd = fwd_idx
                break
            if fwd_idx == 42:
                for bwd_idx in range(1, hv):
                    seq_srh = sheet.cell(row=bwd_idx, column=2).value
                    if seq_srh == key_adhr:
                        fnd = bwd_idx
                        break

        return fnd

    # accept values from choice 1: insert
    def insert_values(self):

        name = uname.get()
        adhr = int(adhaar.get())
        dob = date_val.get()
        mobile = int(phone.get())
        vax_name = vaccine.get().lower()
        dose1 = doseone.get()
        dose2 = dosetwo.get().lower()
        status = stat.get()

        details = [name, adhr, dob, mobile, vax_name, dose1, dose2, status]
        # print(details)
        rec = Hashing()
        rec.Hash_Value(adhr)
        rec.Insert(adhr, details)

    # choice 2: Search
    def look():

        global screen2

        screen2 = Toplevel(screen)
        screen2.title("VAXINATOR - search")
        screen2.geometry("400x200")

        global adhrsrch
        adhrsrch = int

        h = Hashing()

        Label(screen2, height=2, width=50, font=('tahoma', 10),
              text='Enter the Aadhaar Number of Record to be Searched: ').pack()
        adhrsrch = Entry(screen2, text=adhrsrch, relief='solid')
        adhrsrch.pack()
        Button(screen2, bg='#ebf6f9', fg='blue', activebackground='blue', activeforeground='white', text="Search",
               width=10, height=1, command=h.search_values).pack()

    # display record from choice 2: Search
    def search_values(self):

        global screen3

        screen3 = Toplevel(screen)

        screen3.geometry("600x600")

        rec = Hashing()
        uid_val = int(adhrsrch.get())
        rec.Hash_Value(uid_val)
        rec.Search(uid_val)

    # choice 3: Display
    def view():

        global screen4
        screen4 = Toplevel(screen)
        screen4.title("VAXINATOR - all records")
        screen4.geometry("400x600")

        rec = Hashing()
        rec.Display()

    # choice 4: Update
    def change():

        global screen5
        screen5 = Toplevel(screen)
        screen5.title("VAXINATOR - update")
        screen5.geometry("400x200")

        global adhrupd
        adhrupd = int

        h = Hashing()

        Label(screen5, height=2, width=50, font=('tahoma', 10),
              text='Enter the Aadhaar Number of Record to be Updated: ').pack()
        adhrupd = Entry(screen5, text=adhrupd, relief='solid')
        adhrupd.pack()
        Button(screen5, bg='#ebf6f9', fg='orangered', activebackground='lightyellow', activeforeground='orangered',
               text="OK", width=10, height=1, command=h.update_values).pack()

    # update record from choice 4: Update
    def update_values(self):

        global screen6
        screen6 = Toplevel(screen)

        screen6.geometry("600x600")

        global dose2upd
        global statusupd

        dose2upd = StringVar()
        statusupd = StringVar()

        rec = Hashing()

        uid_upd = int(adhrupd.get())
        hv = rec.Hash_Value(uid_upd)

        workbook = load_workbook(filename="record_info.xlsx")
        sheet = workbook.active
        now = date.today()

        fnd = rec.locate(hv, uid_upd)

        if fnd:

            screen5.destroy()
            screen6.title("VAXINATOR - update [Record found at position {} ]".format(fnd))
            sheet = workbook.active
            d1 = sheet["F" + str(fnd)].value
            year1, month1, day1 = map(int, d1.split('-'))
            date1 = datetime.date(year1, month1, day1)
            date2 = sheet["G" + str(fnd)].value
            vac = sheet["E" + str(fnd)].value
            Label(screen6, fg='blue', font=('tahoma', 10), text="DETAILS:\n").pack()
            hd2 = 0
            for data in sheet[str(fnd)]:
                hd2_val = header[hd2]
                Label(screen6, font=('tahoma', 10), text=hd2_val + str(data.value).upper()).pack()
                hd2 += 1

            Label(screen6, text="\n").pack()

            if date2 == "na":
                diff = (now - date1).days
                Label(screen6, font=('tahoma', 10), text="Number of days completed since 1st dose:").pack()
                Label(screen6, fg='blue', font=('tahoma', 12), text=diff).pack()

                if vac == 'covishield':
                    max_days = 84
                    Label(screen6, font=('tahoma', 10), text="\nMinimum Interval for COVISHIELD:").pack()
                    Label(screen6, fg='green', font=('tahoma', 12), text='84 Days').pack()

                else:
                    max_days = 28
                    Label(screen6, font=('tahoma', 10), text="\nMinimum Interval for COVAXIN:").pack()
                    Label(screen6, fg='blue', font=('tahoma', 12), text='28 Days').pack()

                if diff >= max_days:
                    Label(screen6, fg='green', font=('tahoma', 12),
                          text="\nThe person is ELIGIBLE for second dose").pack()
                    Label(screen6, height=2, width=50, font=('tahoma', 10),
                          text='\n\nUpdate dose2 vaccinated date\n(YYYY-MM-DD): ').pack()
                    dose2upd = Entry(screen6, text=dose2upd, relief='solid')
                    dose2upd.pack()
                    Label(screen6, height=2, width=50, font=('tahoma', 10), text='\nUpdate the vaccine status: ').pack()
                    statusupd = Entry(screen6, text=statusupd, relief='solid')
                    statusupd.pack()
                    Button(screen6, bg='#ebf6f9', fg='orangered', activebackground='yellow',
                           activeforeground='orangered', text='UPDATE', width=10, height=1,
                           command=lambda: rec.Update(fnd)).pack()

                else:
                    Label(screen6, fg='red', font=('tahoma', 12),
                          text="\nThe person is NOT eligible for second dose now").pack()
                    visit = date1 + timedelta(days=max_days)
                    Label(screen6, font=('tahoma', 10), text="\nEligible for second dose from:").pack()
                    Label(screen6, fg='blue', font=('tahoma', 12), text=str(visit)).pack()


            else:
                Label(screen6, fg='green', font=('tahoma', 12), text="The Person is Completely Vaccinated!!").pack()

        else:
            messagebox.showerror('Unavailable', 'Record Not Found!')
            screen6.destroy()
            screen5.destroy()
        workbook.save(filename="record_info.xlsx")

    # choice 5: Delete
    def delet():

        global screen7
        screen7 = Toplevel(screen)
        screen7.title("VAXINATOR - delete")
        screen7.geometry("400x200")

        global adhrdel
        adhrdel = int

        h = Hashing()

        Label(screen7, height=2, width=50, font=('tahoma', 10),
              text='Enter the Aadhaar Number of Record to be Deleted: ').pack()
        adhrdel = Entry(screen7, text=adhrdel, relief='solid')
        adhrdel.pack()
        Button(screen7, bg='#ebf6f9', fg='red', activebackground='red', activeforeground='white', text="OK", width=10,
               height=1, command=h.delete_values).pack()

    # delete record from choice 5: Delete
    def delete_values(self):

        global screen8
        screen8 = Toplevel(screen)

        screen8.geometry("500x500")

        uid_num = int(adhrdel.get())

        rec = Hashing()
        hv = rec.Hash_Value(uid_num)
        workbook = load_workbook(filename="record_info.xlsx")
        now = date.today()

        fnd = rec.locate(hv, uid_num)

        if fnd:

            screen7.destroy()
            screen8.title("VAXINATOR - delete [Record found at position {}]".format(fnd))
            sheet = workbook.active
            d1 = sheet["F" + str(fnd)].value
            year1, month1, day1 = map(int, d1.split('-'))
            date1 = datetime.date(year1, month1, day1)
            date2 = sheet["G" + str(fnd)].value
            vac = sheet["E" + str(fnd)].value
            Label(screen8, fg='blue', font=('tahoma', 10), text="DETAILS:\n").pack()
            hd3 = 0
            for data in sheet[str(fnd)]:
                hd3_val = header[hd3]
                Label(screen8, font=('tahoma', 10), text=hd3_val + str(data.value).upper()).pack()
                hd3 += 1

            Label(screen8, text="\n").pack()
            Button(screen8, bg='#ebf6f9', font=('tahoma', 8), fg='red', activebackground='red',
                   activeforeground='white', text='DELETE', width=10, height=1, command=lambda: rec.Remove(fnd)).pack()

        else:
            messagebox.showerror('Unavailable', 'Record Not Found!')
            screen8.destroy()
            screen7.destroy()
        workbook.save(filename="record_info.xlsx")


# home screen
global screen
screen = Tk()  # tkinter.Tk()   used to create the main window
screen.geometry("1080x1080")  # sets the size of the window
screen.title('VAXINATOR - home')  # sets the title for the window
screen.state('zoomed')

header = ['Full Name: ', 'Aadhaar Number: ', 'DOB: ', 'Mobile number: ', 'Vaccine Name: ', 'Dose1 date: ',
          'Dose2 date: ', 'Status: ']

h = Hashing

Label(text="").pack()  # A widget used to display text (or image)on the screen.
Label(bg='lightgreen', font=('tahoma', 20, 'bold', 'underline'), height=2, width=23,
      text='Welcome to VAXINATOR!!').pack()

Label(text="").pack()  # pack() adds widget to the screen
Label(text="").pack()

image = Image.open("syringe.png")  # opens the image
resize_image = image.resize((150, 150))  # resizing the image
img = ImageTk.PhotoImage(resize_image)  # adding resized image to the widget using PhotoImage class
label1 = Label(image=img)
label1.image = img
label1.pack()

Label(height=2, width=30, font=('tahoma', 10), text='Get Vaccinated Now!').pack()
Label(height=2, width=23, font=('tahoma', 14), text='Select an Option:').pack()

# flat, groove, raised, ridge, solid, or sunken

Button(height=2, width=20, bg='#ebf6f9', fg='darkgreen', activebackground='#b5fcb5', activeforeground='#023102',
       font=('tahoma', 10), text='Insert Records', command=h.getvalues).pack()
Label(text="").pack()
Button(height=2, width=20, bg='#ebf6f9', fg='blue', activebackground='#ccccff', activeforeground='blue',
       font=('tahoma', 10), text='Search Records', command=h.look).pack()
Label(text="").pack()
Button(height=2, width=20, bg='#ebf6f9', fg='black', activebackground='whitesmoke', activeforeground='black',
       font=('tahoma', 10), text='Display Records', command=h.view).pack()
Label(text="").pack()
Button(height=2, width=20, bg='#ebf6f9', fg='orangered', activebackground='#ffffa6', activeforeground='orangered',
       font=('tahoma', 10), text='Update Records', command=h.change).pack()
Label(text="").pack()
Button(height=2, width=20, bg='#ebf6f9', fg='red', activebackground='#ffcccc', activeforeground='red',
       font=('tahoma', 10), text='Delete Records', command=h.delet).pack()

screen.mainloop()